
####
'''
This notebook-style script extracts provincial solar PV tables from NEA source pages.

The workflow has four main parts.
First, it scrapes each NEA article page to capture metadata, table image URLs, and whether the page is image-based or an inline HTML table.
Second, it downloads and vertically stitches image tables so split website images become one OCR-ready source image. Inline HTML pages skip this image step.
Third, it extracts raw table rows. Image pages can use OpenAI Vision, direct PaddleOCR reconstruction, or img2table; inline pages use the HTML/text parser. OpenAI Vision calls are guarded by size/cost checks and their raw responses are written to outputs/logs/openai_responses for audit.
Finally, it cleans and reviews the extraction: applies the fixed schema, handles legacy pre-household table layouts, translates province names, converts OCR/text numbers into numeric values, and writes clean CSVs plus review workbooks.

The blocks are split up on purpose. A new user should be able to run them one by one, inspect the outputs at each stage, and stop wherever something looks off. That matters here because NEA tables are similar but not perfectly identical across years and checkpoints, and older tables use different column layouts.

What the steps do:

Steps 1-2 set up the environment: imports, file paths, URL inputs, OCR/OpenAI settings, rerun toggles, fixed schema, province translation map, and helper functions. This is the shared machinery the rest of the script relies on.

Step 3 scrapes the source pages and builds simple manifests, so you can see which pages were found, what metadata was extracted, and how many table images each page contains.

Step 4 downloads the raw image files and stitches them vertically where needed. This creates a consistent input for image-based extraction and gives you a useful visual checkpoint.

Step 5 initializes the local OCR engine only when the selected mode needs it. OpenAI Vision mode skips local OCR initialization.

Step 6 runs one page through the active extraction path and saves the raw output to Excel. For image pages this uses the selected OCR/Vision mode; for inline pages this uses the HTML parser. This is the first real extraction checkpoint.

Step 7 cleans the raw output. It finds the table body, normalizes older sparse/cumulative-first layouts, applies the expected schema, translates province names, and converts value columns from OCR/text strings to numeric form.

Step 8 builds a human-review version of the table. This includes English labels, expected Chinese labels, and the cleaned values in one place so you can visually compare the machine output against the source.

Step 9 adds light workbook formatting to make the review file easier to read.

Step 10 is the batch loop. It can run all pages, rerun only URLs that failed according to outputs/logs/run_summary.csv, or force rerun specific URLs. It updates run_summary.csv after each page so progress and failures are visible during long runs.

'''
####

'''
conda create -n neaocr python=3.10 -y
conda activate neaocr

python -m pip install --upgrade pip
python -m pip install "numpy<2"
python -m pip install paddlepaddle==3.2.0 -i https://www.paddlepaddle.org.cn/packages/stable/cpu/
python -m pip install paddleocr opencv-python pillow beautifulsoup4 requests pandas openpyxl

python -m pip install "img2table[paddle]"

# Required only when STEP6_OCR_MODE = "openai_vision".
# PowerShell:
# $env:OPENAI_API_KEY="your_api_key_here"

python -c "import img2table; print('img2table ok')"
python -c "from img2table.document import Image; print('Image ok')"
python -c "from img2table.ocr import PaddleOCR; print('img2table PaddleOCR wrapper ok')"

'''





#1) Imports, paths, URLs, fixed schema, province map

import re
import os
import sys
import time
import json
import base64
from pathlib import Path
from urllib.parse import urljoin, urlparse

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

import pandas as pd
import requests
import cv2
import numpy as np
from bs4 import BeautifulSoup
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")

from img2table.document import Image as Img2TableImage
from img2table.ocr import PaddleOCR as Img2TablePaddleOCR
from img2table.tables.processing import common as img2table_common

# ============================================================
# USER-EDITABLE PATHS
# ============================================================
BASE_DIR = Path(r"C:\Users\james\Documents\GitHub\GIPT\China NEA")

OUTPUTS_DIR = BASE_DIR / "outputs"
RAW_IMAGES_DIR = OUTPUTS_DIR / "raw_images"
STITCHED_DIR = OUTPUTS_DIR / "stitched"
RAW_TABLES_DIR = OUTPUTS_DIR / "raw_tables_xlsx"
REVIEW_DIR = OUTPUTS_DIR / "review_workbooks"
CLEAN_CSV_DIR = OUTPUTS_DIR / "clean_csv"
LOGS_DIR = OUTPUTS_DIR / "logs"
OPENAI_RESPONSES_DIR = LOGS_DIR / "openai_responses"

for d in [BASE_DIR, OUTPUTS_DIR, RAW_IMAGES_DIR, STITCHED_DIR, RAW_TABLES_DIR, REVIEW_DIR, CLEAN_CSV_DIR, LOGS_DIR, OPENAI_RESPONSES_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ============================================================
# INPUT URLS
# ============================================================
# Fallback/testing list. Used only when URLS_CSV is None.
URLS = [
    "https://www.nea.gov.cn/20260519/6008bb88dc0649e5b5530fd6ad0b117a/c.html",
    "https://www.nea.gov.cn/20260305/4216fa1274bd4b7f8da92090ba3999aa/c.html",
    "https://www.nea.gov.cn/20251112/35126d06a151461882b61d0a2e5706a6/c.html",
    "https://www.nea.gov.cn/20250811/b32802d80ef04148b704e6bc1cd51eb2/c.html",
    "https://www.nea.gov.cn/20250429/b78504d2e8a14b97bdcffb2c501b7393/c.html",
    "https://www.nea.gov.cn/20250221/f04452701c914d51a89d0c0ea6f4acd1/c.html",
    "https://www.nea.gov.cn/2024-11/01/c_1310787081.htm",
    "https://www.nea.gov.cn/2024-07/25/c_1310782757.htm",
    "https://www.nea.gov.cn/2024-05/06/c_1310773741.htm",
]

# Use the full structured URL manifest by default.
# It should contain column 'url' or 'source_page_url'.
URLS_CSV = BASE_DIR / "nea_provincial_pv_source_pages_structured_2016_2026.csv"

# Step 10 toggle. When True, only rerun URLs marked failed in outputs/logs/run_summary.csv.
# Existing successful rows are kept in run_summary.csv and failed rows are replaced
# as they are retried.
RERUN_ONLY_FAILED_FROM_RUN_SUMMARY = False
FORCE_RERUN_URLS = []
ENV_FORCE_RERUN_URLS = os.getenv("NEA_FORCE_RERUN_URLS", "")
if ENV_FORCE_RERUN_URLS.strip():
    FORCE_RERUN_URLS = [
        url.strip()
        for url in re.split(r"[,\s]+", ENV_FORCE_RERUN_URLS)
        if url.strip()
    ]
force_rerun_urls = set(FORCE_RERUN_URLS)
# Example:
# FORCE_RERUN_URLS = [
#     "https://www.nea.gov.cn/2021-04/27/c_139910029.htm",
#     "https://www.nea.gov.cn/2021-10/25/c_1310267679.htm",
# ]

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}

OCR_LANG = "ch"
MIN_CONFIDENCE = 50
TOTAL_LABEL = "总计"
EXPECTED_N_COLS = 9
INLINE_TABLE_PARSER_VERSION = "html-table-v2"
STEP6_OCR_MODE = os.getenv("NEA_OCR_MODE", "openai_vision")  # "openai_vision", "direct_paddle", or "img2table"
OPENAI_VISION_MODEL = "gpt-4.1-mini"
OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
OPENAI_IMAGE_DETAIL = "high"  # "low", "high", or "auto"; high is safer for small table text
OPENAI_MAX_OUTPUT_TOKENS = 5000
OPENAI_MAX_IMAGE_PIXELS = 8_000_000
OPENAI_MAX_IMAGE_FILE_MB = 8
OPENAI_MAX_ESTIMATED_CALL_COST_USD = 0.02
OPENAI_MAX_BATCH_IMAGE_CALLS = 50

OPENAI_MODEL_PRICING = {
    # USD per 1M tokens. Image token estimate follows OpenAI's patch-based
    # sizing for the gpt-4.1-mini 2025-04-14 snapshot family.
    "gpt-4.1-mini": {
        "input_per_1m": 0.40,
        "output_per_1m": 1.60,
        "image_patch_budget": 1536,
        "image_token_multiplier": 1.62,
    },
}

# OCR extraction mode notes:
# - With the current default, STEP6_OCR_MODE = "openai_vision", image tables
#   use the OpenAI helpers below. Local Paddle/img2table helpers are not called.
# - STEP6_OCR_MODE = "direct_paddle" uses Paddle text boxes and custom row/column
#   reconstruction; if that fails, it falls back to img2table.
# - STEP6_OCR_MODE = "img2table" uses img2table directly.
# The local OCR code is kept because it is useful for offline/debug runs, but it
# is intentionally dormant in the default OpenAI Vision workflow.

# img2table can occasionally raise an opaque OpenCV error while looking for
# table titles above detected tables. Returning no contours is consistent with
# img2table's own empty-crop behavior and lets table extraction continue.
_ORIGINAL_GET_CONTOURS_CELL = img2table_common.get_contours_cell


def safe_get_contours_cell(*args, **kwargs):
    try:
        return _ORIGINAL_GET_CONTOURS_CELL(*args, **kwargs)
    except cv2.error as e:
        print(f"WARNING: img2table contour detection skipped after OpenCV error: {e}")
        return []


img2table_common.get_contours_cell = safe_get_contours_cell
if "img2table.tables.processing.text.titles" in sys.modules:
    sys.modules["img2table.tables.processing.text.titles"].get_contours_cell = safe_get_contours_cell

EXPECTED_EN = [
    "province_cn",
    "new_capacity_total_10kw",
    "new_capacity_utility_10kw",
    "new_capacity_distributed_10kw",
    "new_capacity_residential_10kw",
    "cum_capacity_total_10kw",
    "cum_capacity_utility_10kw",
    "cum_capacity_distributed_10kw",
    "cum_capacity_residential_10kw",
]

FRIENDLY_EN = [
    "Province (CN)",
    "New capacity total",
    "New utility-scale",
    "New distributed",
    "New residential",
    "Cumulative total",
    "Cumulative utility-scale",
    "Cumulative distributed",
    "Cumulative residential",
]

PROVINCE_EN_MAP = {
    "总计": "Total",
    "北京": "Beijing",
    "天津": "Tianjin",
    "河北": "Hebei",
    "山西": "Shanxi",
    "内蒙古": "Inner Mongolia",
    "辽宁": "Liaoning",
    "吉林": "Jilin",
    "黑龙江": "Heilongjiang",
    "上海": "Shanghai",
    "江苏": "Jiangsu",
    "浙江": "Zhejiang",
    "安徽": "Anhui",
    "福建": "Fujian",
    "江西": "Jiangxi",
    "山东": "Shandong",
    "河南": "Henan",
    "湖北": "Hubei",
    "湖南": "Hunan",
    "广东": "Guangdong",
    "广西": "Guangxi",
    "海南": "Hainan",
    "重庆": "Chongqing",
    "四川": "Sichuan",
    "贵州": "Guizhou",
    "云南": "Yunnan",
    "西藏": "Tibet",
    "陕西": "Shaanxi",
    "甘肃": "Gansu",
    "青海": "Qinghai",
    "宁夏": "Ningxia",
    "新疆": "Xinjiang",
    "新疆维吾尔自治区": "Xinjiang",
    "新疆兵团": "Xinjiang Production and Construction Corps",
    "新疆生产建设兵团": "Xinjiang Production and Construction Corps",
}


#2 Helper functions


def safe_slug(s):
    s = str(s) if s is not None else "unknown"
    s = re.sub(r"[^\w\-]+", "_", s)
    return s.strip("_")


def load_urls(urls=None, urls_csv=None):
    if urls_csv:
        df = pd.read_csv(urls_csv)
        if "url" in df.columns:
            return df["url"].dropna().tolist()
        if "source_page_url" in df.columns:
            return df["source_page_url"].dropna().tolist()
        raise ValueError("URLS_CSV must contain 'url' or 'source_page_url'")
    return list(urls or [])


def extract_page_metadata(html: str):
    text = BeautifulSoup(html, "html.parser").get_text("\n", strip=True)
    m_date = re.search(r"发布时间：\s*([0-9]{4}-[0-9]{2}-[0-9]{2})", text)
    publish_date = m_date.group(1) if m_date else None
    m_unit = re.search(r"单位[:：]\s*([^\n]+)", text)
    unit = m_unit.group(1).strip() if m_unit else None
    return publish_date, unit


def infer_year_from_title_or_date(title, publish_date):
    if title:
        m = re.search(r"(20\d{2})年", title)
        if m:
            return int(m.group(1))
    if publish_date:
        return int(str(publish_date)[:4])
    return None


def infer_checkpoint(title):
    title = str(title or "")
    if "前三季度" in title:
        return "q3"
    if "上半年" in title:
        return "h1"
    if "一季度" in title:
        return "q1"
    return "annual"


def build_expected_cn_headers(year, checkpoint):
    if year is None:
        year = "UNKNOWN"
    if checkpoint == "q1":
        new_prefix = f"{year}年一季度新增并网容量"
        cum_prefix = f"截至{year}年3月底累计并网容量"
    elif checkpoint == "h1":
        new_prefix = f"{year}年上半年新增并网容量"
        cum_prefix = f"截至{year}年6月底累计并网容量"
    elif checkpoint == "q3":
        new_prefix = f"{year}年前三季度新增并网容量"
        cum_prefix = f"截至{year}年9月底累计并网容量"
    else:
        new_prefix = f"{year}年新增并网容量"
        cum_prefix = f"截至{year}年底累计并网容量"
    return [
        "省(区、市)",
        new_prefix,
        f"{new_prefix}_其中：集中式光伏电站",
        f"{new_prefix}_其中：分布式光伏",
        f"{new_prefix}_其中：户用光伏",
        cum_prefix,
        f"{cum_prefix}_其中：集中式光伏电站",
        f"{cum_prefix}_其中：分布式光伏",
        f"{cum_prefix}_其中：户用光伏",
    ]


def scrape_page(url: str):
    r = requests.get(url, headers=REQUEST_HEADERS, timeout=30)
    r.raise_for_status()
    r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "html.parser")
    h1 = soup.find(["h1", "H1"])
    if h1:
        title = h1.get_text(" ", strip=True)
    else:
        title_tag = soup.find("title")
        title = title_tag.get_text(" ", strip=True) if title_tag else None
    publish_date, unit = extract_page_metadata(r.text)
    year = infer_year_from_title_or_date(title, publish_date)
    checkpoint = infer_checkpoint(title)
    image_urls = []
    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or img.get("zoomfile")
        if not src:
            continue
        full_url = urljoin(url, src)
        if re.search(r"\.(png|jpg|jpeg|webp)(\?.*)?$", full_url, flags=re.I):
            image_urls.append(full_url)
    seen = set()
    image_urls = [x for x in image_urls if not (x in seen or seen.add(x))]
    return {
        "source_page_url": url,
        "title": title,
        "publish_date": publish_date,
        "year": year,
        "checkpoint": checkpoint,
        "unit": unit,
        "n_images": len(image_urls),
        "image_urls": image_urls,
    }


def download_image(url, out_path):
    r = requests.get(url, headers=REQUEST_HEADERS, timeout=30)
    r.raise_for_status()
    out_path.write_bytes(r.content)
    return out_path


def open_image(path):
    return PILImage.open(path).convert("RGB")


def stitch_vertical(image_paths, out_path, bg_color=(255, 255, 255)):
    imgs = [open_image(p) for p in image_paths]
    max_width = max(img.width for img in imgs)
    total_height = sum(img.height for img in imgs)
    canvas = PILImage.new("RGB", (max_width, total_height), color=bg_color)
    y = 0
    for img in imgs:
        canvas.paste(img, (0, y))
        y += img.height
    canvas.save(out_path, quality=95)
    return out_path


def flatten_header(vals):
    out = []
    for v in vals:
        if pd.isna(v):
            continue
        s = str(v).replace("\n", "").strip()
        if not s:
            continue
        if not out or s != out[-1]:
            out.append(s)
    return "".join(out)


def normalize_province_name(x):
    if pd.isna(x):
        return x
    s = str(x).strip()
    s = s.replace(" ", "").replace("\n", "")
    s = s.replace("（", "(").replace("）", ")")
    return s


def to_float(x):
    if pd.isna(x):
        return pd.NA
    s = str(x).strip()
    if s == "":
        return pd.NA
    s = s.replace(",", "").replace("，", "")
    s = s.replace("O", "0").replace("o", "0")
    s = re.sub(r"[^\d\.\-]", "", s)
    if s in {"", ".", "-", "-."}:
        return pd.NA
    try:
        return float(s)
    except Exception:
        return pd.NA


# Local Paddle/img2table OCR helpers.
# These functions are not used when STEP6_OCR_MODE = "openai_vision".
# They are retained for the alternate local modes:
# - direct_paddle: direct PaddleOCR text detection plus custom table assembly
# - img2table: img2table's table extraction wrapper
def box_to_xyxy(box):
    arr = np.asarray(box)
    if arr.ndim == 1 and arr.size == 4:
        x1, y1, x2, y2 = arr.tolist()
    else:
        arr = arr.reshape(-1, 2)
        x1, y1 = arr.min(axis=0).tolist()
        x2, y2 = arr.max(axis=0).tolist()
    return float(x1), float(y1), float(x2), float(y2)


def paddle_result_to_words(ocr_result, min_confidence=MIN_CONFIDENCE):
    rows = []
    min_score = min_confidence / 100
    for text, score, box in zip(
        ocr_result.get("rec_texts", []),
        ocr_result.get("rec_scores", []),
        ocr_result.get("rec_boxes", []),
    ):
        text = str(text).strip()
        score = float(score)
        if not text or score < min_score:
            continue
        x1, y1, x2, y2 = box_to_xyxy(box)
        rows.append({
            "text": text,
            "score": score,
            "x1": x1,
            "y1": y1,
            "x2": x2,
            "y2": y2,
            "xc": (x1 + x2) / 2,
            "yc": (y1 + y2) / 2,
            "height": y2 - y1,
        })
    return pd.DataFrame(rows)


def match_known_province(text):
    clean = normalize_province_name(text)
    if clean in PROVINCE_EN_MAP:
        return clean
    for province in sorted(PROVINCE_EN_MAP, key=len, reverse=True):
        if province in clean:
            return province
    return None


def cluster_1d(values, n_clusters, max_iter=50):
    values = np.asarray(values, dtype=float)
    if len(values) < n_clusters:
        raise ValueError(f"Need at least {n_clusters} OCR positions, got {len(values)}")

    sorted_values = np.sort(values)
    centers = np.array([np.median(chunk) for chunk in np.array_split(sorted_values, n_clusters)])

    for _ in range(max_iter):
        distances = np.abs(values[:, None] - centers[None, :])
        labels = distances.argmin(axis=1)
        next_centers = centers.copy()
        for idx in range(n_clusters):
            members = values[labels == idx]
            if len(members):
                next_centers[idx] = np.median(members)
        if np.allclose(centers, next_centers):
            break
        centers = next_centers

    return np.sort(centers)


def direct_paddle_content(image_path, ocr):
    image = np.array(open_image(image_path))
    paddle_engine = getattr(getattr(ocr, "instance", None), "ocr", None)
    if paddle_engine is not None and hasattr(paddle_engine, "predict"):
        result = paddle_engine.predict(input=[image])
        first = result[0]
        return {
            "rec_texts": list(first["rec_texts"]),
            "rec_scores": list(first["rec_scores"]),
            "rec_boxes": [box.tolist() if hasattr(box, "tolist") else box for box in first["rec_boxes"]],
        }

    doc = Img2TableImage(str(image_path))
    return ocr.content(document=doc)[0]


def extract_raw_table_direct_paddle(image_path, ocr, expected_n_cols=EXPECTED_N_COLS):
    ocr_result = direct_paddle_content(image_path, ocr)
    words = paddle_result_to_words(ocr_result)
    if words.empty:
        raise ValueError("Direct PaddleOCR returned no text")

    words["province_match"] = words["text"].apply(match_known_province)
    row_labels = words[words["province_match"].notna()].sort_values(["yc", "x1"]).copy()
    if row_labels.empty:
        raise ValueError("Direct PaddleOCR could not find any known province/total labels")

    median_height = max(8, float(words["height"].median()))
    row_cluster_tol = max(10, median_height * 1.4)
    row_token_tol = max(12, median_height * 1.6)

    row_anchors = []
    for item in row_labels.itertuples(index=False):
        if row_anchors and abs(item.yc - row_anchors[-1]["yc"]) <= row_cluster_tol:
            row_anchors[-1]["yc_values"].append(item.yc)
            if item.x1 < row_anchors[-1]["x1"]:
                row_anchors[-1]["label"] = item.province_match
                row_anchors[-1]["x1"] = item.x1
        else:
            row_anchors.append({
                "yc": item.yc,
                "yc_values": [item.yc],
                "label": item.province_match,
                "x1": item.x1,
            })

    for anchor in row_anchors:
        anchor["yc"] = float(np.median(anchor["yc_values"]))

    body_tokens = []
    for anchor in row_anchors:
        row_words = words[words["yc"].sub(anchor["yc"]).abs() <= row_token_tol].copy()
        row_words["row_yc"] = anchor["yc"]
        row_words["row_label"] = anchor["label"]
        body_tokens.append(row_words)

    body_tokens = pd.concat(body_tokens, ignore_index=True)
    column_centers = cluster_1d(body_tokens["xc"].tolist(), expected_n_cols)

    output_rows = []
    seen_y = set()
    for anchor in row_anchors:
        row_key = round(anchor["yc"], 1)
        if row_key in seen_y:
            continue
        seen_y.add(row_key)

        row_words = body_tokens[body_tokens["row_yc"].eq(anchor["yc"])].copy()
        cells = [[] for _ in range(expected_n_cols)]
        for token in row_words.sort_values(["x1", "y1"]).itertuples(index=False):
            col_idx = int(np.argmin(np.abs(column_centers - token.xc)))
            cells[col_idx].append(token.text)

        values = ["".join(parts).strip() for parts in cells]
        if not values[0]:
            values[0] = anchor["label"]
        if len(values) == expected_n_cols:
            output_rows.append(values)

    raw_df = pd.DataFrame(output_rows)
    if raw_df.empty:
        raise ValueError("Direct PaddleOCR did not reconstruct any table rows")
    if not raw_df.iloc[:, 0].astype(str).str.strip().eq(TOTAL_LABEL).any():
        raise ValueError(f"Direct PaddleOCR did not reconstruct a '{TOTAL_LABEL}' row")
    return raw_df


def extract_raw_table_img2table(image_path, ocr):
    doc = Img2TableImage(str(image_path))
    tables = doc.extract_tables(
        ocr=ocr,
        implicit_rows=True,
        implicit_columns=True,
        borderless_tables=False,
        min_confidence=MIN_CONFIDENCE,
    )
    print("n_tables:", len(tables))
    if len(tables) == 0:
        raise ValueError("No tables returned by OCR")
    return tables[0].df.copy()


# OpenAI Vision extraction helpers.
# These are the active image-table extraction helpers when the default
# STEP6_OCR_MODE = "openai_vision" is used.
def image_to_data_url(image_path):
    suffix = Path(image_path).suffix.lower()
    media_type = "image/png" if suffix == ".png" else "image/jpeg"
    encoded = base64.b64encode(Path(image_path).read_bytes()).decode("ascii")
    return f"data:{media_type};base64,{encoded}"


def estimate_openai_vision_call_cost(image_path, prompt, model, max_output_tokens):
    pricing = OPENAI_MODEL_PRICING.get(model)
    if pricing is None:
        raise ValueError(
            f"No local OpenAI pricing guard for model '{model}'. "
            "Add it to OPENAI_MODEL_PRICING before making API calls."
        )

    image_path = Path(image_path)
    with PILImage.open(image_path) as img:
        width, height = img.size

    image_pixels = width * height
    image_file_mb = image_path.stat().st_size / (1024 * 1024)
    patch_count = int(np.ceil(width / 32) * np.ceil(height / 32))
    resized_patch_count = min(patch_count, pricing["image_patch_budget"])
    image_tokens = int(np.ceil(resized_patch_count * pricing["image_token_multiplier"]))
    prompt_tokens = max(1, int(np.ceil(len(prompt) / 4)))
    estimated_input_tokens = image_tokens + prompt_tokens

    estimated_input_cost = estimated_input_tokens / 1_000_000 * pricing["input_per_1m"]
    estimated_output_cost = max_output_tokens / 1_000_000 * pricing["output_per_1m"]

    return {
        "width": width,
        "height": height,
        "image_pixels": image_pixels,
        "image_file_mb": image_file_mb,
        "image_tokens": image_tokens,
        "prompt_tokens": prompt_tokens,
        "estimated_input_tokens": estimated_input_tokens,
        "max_output_tokens": max_output_tokens,
        "estimated_max_cost_usd": estimated_input_cost + estimated_output_cost,
    }


def guard_openai_vision_call(image_path, prompt):
    estimate = estimate_openai_vision_call_cost(
        image_path=image_path,
        prompt=prompt,
        model=OPENAI_VISION_MODEL,
        max_output_tokens=OPENAI_MAX_OUTPUT_TOKENS,
    )

    print(
        "OpenAI estimated max cost: "
        f"${estimate['estimated_max_cost_usd']:.4f} "
        f"({estimate['width']}x{estimate['height']}, "
        f"{estimate['image_file_mb']:.2f} MB, "
        f"~{estimate['estimated_input_tokens']} input tokens, "
        f"max {estimate['max_output_tokens']} output tokens)"
    )

    if estimate["image_pixels"] > OPENAI_MAX_IMAGE_PIXELS:
        raise ValueError(
            "OpenAI call blocked: image is too large "
            f"({estimate['image_pixels']:,} pixels > {OPENAI_MAX_IMAGE_PIXELS:,})."
        )
    if estimate["image_file_mb"] > OPENAI_MAX_IMAGE_FILE_MB:
        raise ValueError(
            "OpenAI call blocked: image file is too large "
            f"({estimate['image_file_mb']:.2f} MB > {OPENAI_MAX_IMAGE_FILE_MB} MB)."
        )
    if estimate["estimated_max_cost_usd"] > OPENAI_MAX_ESTIMATED_CALL_COST_USD:
        raise ValueError(
            "OpenAI call blocked: estimated max cost "
            f"${estimate['estimated_max_cost_usd']:.4f} exceeds "
            f"${OPENAI_MAX_ESTIMATED_CALL_COST_USD:.4f}."
        )

    return estimate


def response_output_text(response_json):
    chunks = []

    def walk(obj):
        if isinstance(obj, dict):
            if obj.get("type") in {"output_text", "text"} and isinstance(obj.get("text"), str):
                chunks.append(obj["text"])
            for value in obj.values():
                walk(value)
        elif isinstance(obj, list):
            for value in obj:
                walk(value)

    walk(response_json.get("output", []))
    if not chunks and isinstance(response_json.get("output_text"), str):
        chunks.append(response_json["output_text"])
    return "\n".join(chunks).strip()


def parse_json_object_or_array(text):
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start_candidates = [idx for idx in [text.find("{"), text.find("[")] if idx != -1]
        if not start_candidates:
            raise
        start = min(start_candidates)
        end = max(text.rfind("}"), text.rfind("]"))
        return json.loads(text[start:end + 1])


def validate_llm_table_rows(payload, expected_n_cols=EXPECTED_N_COLS):
    rows = payload.get("rows") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise ValueError("OpenAI response did not contain a JSON list of rows")

    cleaned_rows = []
    for row in rows:
        if not isinstance(row, list):
            raise ValueError("OpenAI response included a non-list row")
        row = ["" if v is None else str(v).strip() for v in row]
        if len(row) != expected_n_cols:
            raise ValueError(f"OpenAI row has {len(row)} columns; expected {expected_n_cols}: {row}")
        cleaned_rows.append(row)

    raw_df = pd.DataFrame(cleaned_rows)
    if raw_df.empty:
        raise ValueError("OpenAI response returned no table rows")
    if not raw_df.iloc[:, 0].astype(str).str.strip().eq(TOTAL_LABEL).any():
        raise ValueError(f"OpenAI response did not include a '{TOTAL_LABEL}' row")
    return raw_df


def extract_raw_table_openai_vision(image_path):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("Set OPENAI_API_KEY before using STEP6_OCR_MODE='openai_vision'")

    prompt = f"""
Extract the body rows from this Chinese NEA solar PV table image.

Return only valid JSON, with this exact shape:
{{"rows": [[...], [...]]}}

Rules:
- Include table body rows only, starting with "{TOTAL_LABEL}".
- Exclude title text, unit text, and header rows.
- Each row must have exactly {EXPECTED_N_COLS} string values.
- Keep Chinese province names as Chinese.
- Keep numeric values exactly as shown, including decimals.
- Do not translate, summarize, add markdown, or add comments.
- The column order is:
  1. 省(区、市)
  2. 新增并网容量 total
  3. 新增并网容量 其中：集中式光伏电站
  4. 新增并网容量 其中：分布式光伏
  5. 新增并网容量 其中：户用光伏
  6. 累计并网容量 total
  7. 累计并网容量 其中：集中式光伏电站
  8. 累计并网容量 其中：分布式光伏
  9. 累计并网容量 其中：户用光伏
""".strip()

    guard_openai_vision_call(image_path, prompt)

    payload = {
        "model": OPENAI_VISION_MODEL,
        "input": [{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {
                    "type": "input_image",
                    "image_url": image_to_data_url(image_path),
                    "detail": OPENAI_IMAGE_DETAIL,
                },
            ],
        }],
        "max_output_tokens": OPENAI_MAX_OUTPUT_TOKENS,
    }

    t0 = time.time()
    response = requests.post(
        OPENAI_RESPONSES_URL,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=120,
    )
    print("OpenAI vision API sec:", round(time.time() - t0, 2))
    if response.status_code >= 400:
        raise ValueError(f"OpenAI API error {response.status_code}: {response.text[:1000]}")

    response_json = response.json()
    raw_text = response_output_text(response_json)
    if not raw_text:
        raise ValueError("OpenAI response did not include output text")

    log_path = OPENAI_RESPONSES_DIR / f"{Path(image_path).stem}_openai_vision_response.json"
    log_path.write_text(
        json.dumps({"response": response_json, "output_text": raw_text}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print("saved OpenAI raw response:", log_path)

    parsed = parse_json_object_or_array(raw_text)
    raw_df = validate_llm_table_rows(parsed)
    print("OpenAI vision rows:", len(raw_df))
    return raw_df


def extract_raw_table_from_image(image_path, ocr, mode=STEP6_OCR_MODE):
    # Single routing point for image-based pages. Inline HTML pages bypass this
    # and go through extract_inline_table_from_html().
    if mode == "openai_vision":
        print("using OpenAI vision table extraction")
        return extract_raw_table_openai_vision(image_path)
    if mode == "direct_paddle":
        try:
            print("using direct PaddleOCR table reconstruction")
            raw_df = extract_raw_table_direct_paddle(image_path, ocr)
            print("direct PaddleOCR rows:", len(raw_df))
            return raw_df
        except Exception as e:
            print("direct PaddleOCR failed; falling back to img2table:", e)
            return extract_raw_table_img2table(image_path, ocr)
    if mode == "img2table":
        print("using img2table extraction")
        return extract_raw_table_img2table(image_path, ocr)
    raise ValueError(f"Unknown STEP6_OCR_MODE: {mode}")


def find_total_row_and_start_col(raw_df, total_label=TOTAL_LABEL, max_search_cols=4):
    for col_idx in range(min(max_search_cols, raw_df.shape[1])):
        ser = raw_df[col_idx].astype(str).str.strip()
        matches = raw_df.index[ser.eq(total_label)]
        if len(matches) > 0:
            return int(matches[0]), col_idx
    raise ValueError(f"Could not find total row labeled '{total_label}'.")


def autosize_worksheet(ws, max_width=35):
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = "" if row[0] is None else str(row[0])
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def looks_like_pandas_excel_index_header(raw):
    if raw.shape[0] < 2 or raw.shape[1] < EXPECTED_N_COLS + 1:
        return False

    header_vals = pd.to_numeric(
        raw.iloc[0, 1:EXPECTED_N_COLS + 1],
        errors="coerce",
    )
    if header_vals.isna().any():
        return False
    if header_vals.astype(int).tolist() != list(range(EXPECTED_N_COLS)):
        return False

    index_vals = pd.to_numeric(raw.iloc[1:min(len(raw), 8), 0], errors="coerce")
    if index_vals.isna().any():
        return False
    return index_vals.astype(int).tolist() == list(range(len(index_vals)))


def normalize_raw_excel_frame(raw):
    raw = raw.copy()
    if looks_like_pandas_excel_index_header(raw):
        raw = raw.iloc[1:, 1:].reset_index(drop=True)
    while raw.shape[1] < EXPECTED_N_COLS:
        raw[raw.shape[1]] = pd.NA
    return raw


def normalize_legacy_sparse_capacity_frame(raw, year=None):
    raw = normalize_raw_excel_frame(raw)
    try:
        year = int(float(year))
    except Exception:
        year = None

    if year is None or year > 2021 or raw.empty:
        return raw

    numeric = raw.copy()
    for col in range(1, min(raw.shape[1], EXPECTED_N_COLS)):
        numeric[col] = raw[col].apply(to_float)

    total_rows = numeric[numeric[0].astype(str).str.strip().eq(TOTAL_LABEL)]
    if total_rows.empty:
        return raw

    total = total_rows.iloc[0]
    sparse_legacy = (
        raw.shape[1] >= EXPECTED_N_COLS
        and pd.notna(total[1])
        and numeric[2].isna().mean() >= 0.8
        and numeric[3].isna().mean() >= 0.8
        and numeric[4].isna().mean() >= 0.8
        and pd.notna(total[5])
        and pd.notna(total[6])
        and pd.notna(total[7])
        and float(total[1]) > float(total[5])
    )
    if not sparse_legacy:
        return raw

    fixed = pd.DataFrame(index=raw.index, columns=range(EXPECTED_N_COLS))
    fixed[0] = raw[0]
    fixed[1] = numeric[6]  # new total
    fixed[2] = numeric[7]  # new utility/station
    fixed[3] = numeric[6] - numeric[7]  # derived new distributed
    fixed[4] = pd.NA
    fixed[5] = numeric[1]  # cumulative total
    fixed[6] = numeric[5]  # cumulative utility/station
    fixed[7] = numeric[1] - numeric[5]  # derived cumulative distributed
    fixed[8] = pd.NA
    print("normalized legacy sparse cumulative-first capacity frame")
    return fixed


def repair_legacy_no_household_columns(clean, year, checkpoint):
    try:
        year = int(float(year))
    except Exception:
        return clean
    checkpoint = str(checkpoint or "").strip().lower()

    if year > 2021:
        return clean

    total_rows = clean[clean["province_cn"].astype(str).str.strip().eq(TOTAL_LABEL)]
    if total_rows.empty:
        return clean

    total = total_rows.iloc[0]

    if year <= 2020:
        # Damaged legacy layout A:
        # cum total landed in new total; cum utility landed in cum total;
        # new total landed in cum utility; new utility landed in cum distributed.
        sparse_new_cols = (
            clean["new_capacity_utility_10kw"].isna().mean() >= 0.8
            and clean["new_capacity_distributed_10kw"].isna().mean() >= 0.8
        )
        if (
            sparse_new_cols
            and pd.notna(total["new_capacity_total_10kw"])
            and pd.notna(total["cum_capacity_total_10kw"])
            and pd.notna(total["cum_capacity_utility_10kw"])
            and total["new_capacity_total_10kw"] > total["cum_capacity_total_10kw"]
        ):
            old_cum_total = clean["new_capacity_total_10kw"].copy()
            old_cum_utility = clean["cum_capacity_total_10kw"].copy()
            old_new_total = clean["cum_capacity_utility_10kw"].copy()
            old_new_utility = clean["cum_capacity_distributed_10kw"].copy()

            clean["new_capacity_total_10kw"] = old_new_total
            clean["new_capacity_utility_10kw"] = old_new_utility
            clean["new_capacity_distributed_10kw"] = old_new_total - old_new_utility
            clean["new_capacity_residential_10kw"] = pd.NA
            clean["cum_capacity_total_10kw"] = old_cum_total
            clean["cum_capacity_utility_10kw"] = old_cum_utility
            clean["cum_capacity_distributed_10kw"] = old_cum_total - old_cum_utility
            clean["cum_capacity_residential_10kw"] = pd.NA
            print("repaired legacy cumulative-first capacity layout")
            return clean

        # Damaged legacy layout B:
        # cum total/utility stayed in the new total/utility slots, while
        # new total/utility followed them before curtailment columns.
        if (
            pd.notna(total["new_capacity_total_10kw"])
            and pd.notna(total["new_capacity_utility_10kw"])
            and pd.notna(total["new_capacity_distributed_10kw"])
            and pd.notna(total["cum_capacity_total_10kw"])
            and total["new_capacity_total_10kw"] > total["cum_capacity_total_10kw"] * 5
            and total["new_capacity_utility_10kw"] > total["cum_capacity_total_10kw"] * 5
        ):
            old_cum_total = clean["new_capacity_total_10kw"].copy()
            old_cum_utility = clean["new_capacity_utility_10kw"].copy()
            old_new_total = clean["new_capacity_distributed_10kw"].copy()
            old_new_utility = clean["cum_capacity_total_10kw"].copy()

            clean["new_capacity_total_10kw"] = old_new_total
            clean["new_capacity_utility_10kw"] = old_new_utility
            clean["new_capacity_distributed_10kw"] = old_new_total - old_new_utility
            clean["new_capacity_residential_10kw"] = pd.NA
            clean["cum_capacity_total_10kw"] = old_cum_total
            clean["cum_capacity_utility_10kw"] = old_cum_utility
            clean["cum_capacity_distributed_10kw"] = old_cum_total - old_cum_utility
            clean["cum_capacity_residential_10kw"] = pd.NA
            print("repaired legacy cumulative-first capacity layout")
            return clean

        # Damaged legacy layout C, seen on image-extracted 2019 Q1:
        # new total/utility are correct, then cum total/utility are next.
        if (
            pd.notna(total["new_capacity_total_10kw"])
            and pd.notna(total["new_capacity_distributed_10kw"])
            and pd.notna(total["new_capacity_residential_10kw"])
            and pd.notna(total["cum_capacity_total_10kw"])
            and total["cum_capacity_total_10kw"] < total["new_capacity_total_10kw"]
            and total["new_capacity_distributed_10kw"] > total["new_capacity_total_10kw"] * 5
            and total["new_capacity_residential_10kw"] > total["new_capacity_total_10kw"] * 5
        ):
            old_new_total = clean["new_capacity_total_10kw"].copy()
            old_new_utility = clean["new_capacity_utility_10kw"].copy()
            old_cum_total = clean["new_capacity_distributed_10kw"].copy()
            old_cum_utility = clean["new_capacity_residential_10kw"].copy()

            clean["new_capacity_total_10kw"] = old_new_total
            clean["new_capacity_utility_10kw"] = old_new_utility
            clean["new_capacity_distributed_10kw"] = old_new_total - old_new_utility
            clean["new_capacity_residential_10kw"] = pd.NA
            clean["cum_capacity_total_10kw"] = old_cum_total
            clean["cum_capacity_utility_10kw"] = old_cum_utility
            clean["cum_capacity_distributed_10kw"] = old_cum_total - old_cum_utility
            clean["cum_capacity_residential_10kw"] = pd.NA
            print("repaired legacy cumulative-first capacity layout")
            return clean

    if checkpoint != "q3":
        return clean

    cum_distributed = clean["cum_capacity_distributed_10kw"]
    cum_residential = clean["cum_capacity_residential_10kw"]
    distributed_missing = cum_distributed.isna() | cum_distributed.eq(0)
    residential_present = cum_residential.notna() & ~cum_residential.eq(0)

    if distributed_missing.mean() >= 0.8 and residential_present.mean() >= 0.8:
        clean["cum_capacity_distributed_10kw"] = clean["cum_capacity_residential_10kw"]
        clean["cum_capacity_residential_10kw"] = pd.NA

    return clean


def save_run_summary(run_rows, out_path=LOGS_DIR / "run_summary.csv"):
    run_summary_df = pd.DataFrame(run_rows)
    run_summary_df.to_csv(out_path, index=False, encoding="utf-8-sig")
    return run_summary_df


def load_previous_run_summary(out_path=LOGS_DIR / "run_summary.csv"):
    if not out_path.exists():
        raise FileNotFoundError(
            f"{out_path} does not exist. Run Step 10 once with "
            "RERUN_ONLY_FAILED_FROM_RUN_SUMMARY=False first."
        )
    df = pd.read_csv(out_path)
    required_cols = {"source_page_url", "status"}
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(f"{out_path} is missing columns: {sorted(missing)}")
    return df


def upsert_run_row(run_rows, row):
    url = row["source_page_url"]
    for idx, existing in enumerate(run_rows):
        if existing.get("source_page_url") == url:
            run_rows[idx] = row
            return
    run_rows.append(row)


def is_numeric_cell(x):
    s = "" if x is None else str(x).strip()
    s = normalize_numeric_cell_text(s)
    return s not in {"", ".", "-", "-."}


def normalize_numeric_cell_text(x):
    s = "" if x is None else str(x).strip()
    s = re.sub(r"\s+", "", s)
    return re.sub(r"[^\d\.\-]", "", s)


def subtract_numeric_text(left, right, require_nonnegative=False):
    left_value = to_float(normalize_numeric_cell_text(left))
    right_value = to_float(normalize_numeric_cell_text(right))
    if pd.isna(left_value) or pd.isna(right_value):
        return ""
    difference = left_value - right_value
    if require_nonnegative and difference < 0:
        return ""
    return f"{difference:g}"


def infer_inline_table_layout(table):
    header_text = []
    for tr in table.find_all("tr")[:8]:
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if not cells:
            continue
        if match_known_province(cells[0]) is not None:
            break
        # NEA pages often wrap the real table inside an outer table whose
        # first row contains the whole article text. Ignore that wrapper row
        # so the layout decision is based on the actual table header.
        if len(cells) <= 12:
            header_text.append(" ".join(cells))
    header_text = " ".join(header_text)

    cumulative_pos = header_text.find("累计")
    new_pos = header_text.find("新增")
    if cumulative_pos != -1 and new_pos != -1 and cumulative_pos < new_pos:
        return "cumulative_first_total_utility"
    return "standard"


def normalize_inline_table_row(cells, table_layout="standard"):
    cells = ["" if c is None else str(c).strip() for c in cells]
    if not cells:
        return None

    label = normalize_province_name(cells[0])
    if match_known_province(label) is None:
        return None

    vals = cells[1:]
    if table_layout == "cumulative_first_total_utility" and len(vals) >= 4:
        # Older NEA tables list cumulative capacity first, then new capacity.
        # Some pages append curtailment columns after these four capacity values.
        cumulative_total = normalize_numeric_cell_text(vals[0])
        cumulative_utility = normalize_numeric_cell_text(vals[1])
        new_total = normalize_numeric_cell_text(vals[2]) or "0.0"
        new_utility = normalize_numeric_cell_text(vals[3]) or "0.0"
        vals = [
            new_total,
            new_utility,
            subtract_numeric_text(new_total, new_utility, require_nonnegative=True),
            "",
            cumulative_total,
            cumulative_utility,
            subtract_numeric_text(cumulative_total, cumulative_utility),
            "",
        ]
        required_numeric_idxs = [0, 1, 4, 5, 6]
    elif len(vals) == 4:
        vals = [
            vals[0],
            "",
            "",
            "",
            vals[1],
            vals[2],
            vals[3],
            "",
        ]
        required_numeric_idxs = [0, 4, 5, 6]
        vals = [normalize_numeric_cell_text(v) for v in vals]
    elif len(vals) == 6:
        vals = vals[:3] + [""] + vals[3:] + [""]
        required_numeric_idxs = list(range(8))
        vals = [normalize_numeric_cell_text(v) or "0.0" for v in vals]
    elif len(vals) == 7:
        # Older inline NEA pages publish new household capacity, but not
        # cumulative household capacity. Keep the fixed 9-column schema.
        required_numeric_idxs = list(range(7))
        vals = [normalize_numeric_cell_text(v) for v in vals] + [""]
    else:
        required_numeric_idxs = list(range(8))
        vals = [normalize_numeric_cell_text(v) or "0.0" for v in vals]

    if len(vals) != 8:
        return None
    if not all(is_numeric_cell(vals[idx]) for idx in required_numeric_idxs):
        return None
    if vals[7] and not is_numeric_cell(vals[7]):
        return None

    return [label] + vals


def extract_inline_table_from_html(url: str):
    r = requests.get(url, headers=REQUEST_HEADERS, timeout=30)
    r.raise_for_status()
    r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "html.parser")

    best_rows = []
    best_layout = None
    for table in soup.find_all("table"):
        table_layout = infer_inline_table_layout(table)
        table_rows = []
        for tr in table.find_all("tr"):
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
            row = normalize_inline_table_row(cells, table_layout=table_layout)
            if row is not None:
                table_rows.append(row)
        is_better = len(table_rows) > len(best_rows)
        is_cumulative_tie = (
            len(table_rows) == len(best_rows)
            and table_layout == "cumulative_first_total_utility"
            and best_layout != "cumulative_first_total_utility"
        )
        if is_better or is_cumulative_tie:
            best_rows = table_rows
            best_layout = table_layout

    if best_rows:
        print(
            f"{INLINE_TABLE_PARSER_VERSION}: parsed {len(best_rows)} rows "
            f"from HTML table ({best_layout})"
        )
        return pd.DataFrame(best_rows)

    lines = [x.strip() for x in soup.get_text("\n").splitlines() if x.strip()]
    start_idx = lines.index("总计")
    rows = []
    i = start_idx
    while i < len(lines):
        label = lines[i]
        if match_known_province(label) is None:
            break

        next_label_idx = None
        for j in range(i + 1, min(i + 12, len(lines))):
            if match_known_province(lines[j]) is not None:
                next_label_idx = j
                break
        vals = lines[i + 1:next_label_idx] if next_label_idx else lines[i + 1:i + 9]

        row = normalize_inline_table_row([label] + vals)
        if row is None:
            break
        rows.append(row)
        if next_label_idx is None:
            break
        i = next_label_idx
    print(f"{INLINE_TABLE_PARSER_VERSION}: parsed {len(rows)} rows from page text fallback")
    return pd.DataFrame(rows)


#3) Scrape all source pages and build refs

urls = load_urls(URLS, URLS_CSV)
print(f"{len(urls)} urls loaded")

page_results = [scrape_page(url) for url in urls]

summary_df = pd.DataFrame([
    {
        "source_page_url": r["source_page_url"],
        "title": r["title"],
        "publish_date": r["publish_date"],
        "year": r["year"],
        "checkpoint": r["checkpoint"],
        "unit": r["unit"],
        "n_images": r["n_images"],
    }
    for r in page_results
])

summary_df["page_type"] = summary_df["n_images"].apply(
    lambda x: "image_table" if x > 0 else "inline_html_table"
)
print(summary_df[["publish_date", "year", "checkpoint", "n_images", "page_type", "title"]])

images_df = pd.DataFrame([
    {
        "source_page_url": r["source_page_url"],
        "title": r["title"],
        "publish_date": r["publish_date"],
        "year": r["year"],
        "checkpoint": r["checkpoint"],
        "unit": r["unit"],
        "image_order": i + 1,
        "image_url": img_url,
    }
    for r in page_results
    for i, img_url in enumerate(r["image_urls"])
])

summary_df.to_csv(LOGS_DIR / "page_summary_scraped.csv", index=False, encoding="utf-8-sig")
images_df.to_csv(LOGS_DIR / "image_manifest_scraped.csv", index=False, encoding="utf-8-sig")

print(summary_df)
print(images_df.head())


#4) Download raw images and stitch them
# Step 4 only applies to pages where n_images > 0.
# Inline HTML pages will not appear in images_df and will be handled later.

download_manifest_rows = []
download_images_df = images_df
if force_rerun_urls:
    download_images_df = images_df[images_df["source_page_url"].isin(force_rerun_urls)].copy()

for page_url, g in download_images_df.sort_values(["source_page_url", "image_order"]).groupby("source_page_url", sort=False):
    first = g.iloc[0]
    page_id = f"{first['publish_date']}_{safe_slug(first['title'])[:80]}"
    page_slug = safe_slug(page_id)
    page_raw_dir = RAW_IMAGES_DIR / page_slug
    page_raw_dir.mkdir(parents=True, exist_ok=True)
    local_paths = []
    for row in g.itertuples(index=False):
        ext = Path(urlparse(row.image_url).path).suffix.lower() or ".jpg"
        fname = f"{int(row.image_order):02d}{ext}"
        out_path = page_raw_dir / fname
        download_image(row.image_url, out_path)
        local_paths.append(out_path)
        download_manifest_rows.append({
            "source_page_url": row.source_page_url,
            "title": row.title,
            "publish_date": row.publish_date,
            "year": row.year,
            "checkpoint": row.checkpoint,
            "image_order": row.image_order,
            "image_url": row.image_url,
            "local_path": str(out_path),
        })
    stitched_path = STITCHED_DIR / f"{page_slug}_stitched.jpg"
    stitch_vertical(local_paths, stitched_path)
    print("saved stitched image:", stitched_path)

download_manifest_df = pd.DataFrame(download_manifest_rows)
if force_rerun_urls and (LOGS_DIR / "download_manifest.csv").exists():
    previous_download_manifest = pd.read_csv(LOGS_DIR / "download_manifest.csv")
    previous_download_manifest = previous_download_manifest[
        ~previous_download_manifest["source_page_url"].isin(force_rerun_urls)
    ]
    download_manifest_df = pd.concat(
        [previous_download_manifest, download_manifest_df],
        ignore_index=True,
    )
download_manifest_df.to_csv(LOGS_DIR / "download_manifest.csv", index=False, encoding="utf-8-sig")


#5) initialize ocr (only needed for local Paddle/img2table modes)
if STEP6_OCR_MODE in {"direct_paddle", "img2table"}:
    ocr = Img2TablePaddleOCR(lang=OCR_LANG)
    print("OCR initialized")
else:
    ocr = None
    print(f"Skipping local OCR initialization for STEP6_OCR_MODE={STEP6_OCR_MODE}")


#6) Pick one page and run OCR to raw xlsx
PAGE_IDX = -3   # change to test pages one by one
if force_rerun_urls:
    PAGE_IDX = next(
        (
            idx
            for idx, page_meta in enumerate(page_results)
            if page_meta["source_page_url"] in force_rerun_urls
        ),
        PAGE_IDX,
    )

page_meta = page_results[PAGE_IDX]
print(page_meta["title"])
print(page_meta["publish_date"], page_meta["checkpoint"], page_meta["n_images"])

page_id = f"{page_meta['publish_date']}_{safe_slug(page_meta['title'])[:80]}"
page_slug = safe_slug(page_id)

if page_meta["n_images"] > 0:
    stitched_path = STITCHED_DIR / f"{page_slug}_stitched.jpg"
    print(stitched_path)
    print(stitched_path.exists())
    t0 = time.time()
    raw_df = extract_raw_table_from_image(stitched_path, ocr)
    print("elapsed sec:", round(time.time() - t0, 2))
else:
    t0 = time.time()
    raw_df = extract_inline_table_from_html(page_meta["source_page_url"])
    raw_df = normalize_legacy_sparse_capacity_frame(raw_df, page_meta["year"])
    print("inline html rows:", len(raw_df))
    print("elapsed sec:", round(time.time() - t0, 2))

raw_table_xlsx = RAW_TABLES_DIR / f"{page_slug}_raw.xlsx"
raw_df.to_excel(raw_table_xlsx, index=False, header=False)

print("saved raw table:", raw_table_xlsx)
print(raw_df.head())




#7) Clean the raw xlsx and build review workbook

raw = normalize_raw_excel_frame(pd.read_excel(raw_table_xlsx, header=None))

year = page_meta["year"]
checkpoint = page_meta["checkpoint"]
expected_cn = build_expected_cn_headers(year, checkpoint)

if page_meta["n_images"] > 0:
    data_start_idx, start_col = find_total_row_and_start_col(raw, total_label=TOTAL_LABEL, max_search_cols=4)
    header_start = max(0, data_start_idx - 5)
    header_rows = raw.iloc[header_start:data_start_idx, start_col:start_col + EXPECTED_N_COLS].copy()
    ocr_header_preview = pd.DataFrame({
        "col_position": range(1, EXPECTED_N_COLS + 1),
        "ocr_header_flat": [
            flatten_header(header_rows[c].tolist()) if c in header_rows.columns else ""
            for c in range(start_col, start_col + EXPECTED_N_COLS)
        ],
        "expected_cn": expected_cn,
        "expected_en": EXPECTED_EN,
        "friendly_en": FRIENDLY_EN,
    })
    body = raw.iloc[data_start_idx:, start_col:start_col + EXPECTED_N_COLS].copy()
    body.columns = expected_cn
else:
    # inline HTML pages already come out as 9 clean columns
    raw = normalize_raw_excel_frame(raw)
    body = raw.iloc[:, :EXPECTED_N_COLS].copy()
    body.columns = expected_cn
    # no OCR header to inspect here, so just use expected labels as the preview
    ocr_header_preview = pd.DataFrame({
        "col_position": range(1, EXPECTED_N_COLS + 1),
        "ocr_header_flat": expected_cn,
        "expected_cn": expected_cn,
        "expected_en": EXPECTED_EN,
        "friendly_en": FRIENDLY_EN,
    })

body["省(区、市)"] = body["省(区、市)"].apply(normalize_province_name)
body["province_en"] = body["省(区、市)"].map(PROVINCE_EN_MAP)

for c in expected_cn[1:]:
    body[c] = body[c].apply(to_float)

clean = body.rename(columns=dict(zip(expected_cn, EXPECTED_EN)))
clean = repair_legacy_no_household_columns(clean, year, checkpoint)
clean["province_translation_missing"] = clean["province_en"].isna()

clean = clean[
    ["province_cn", "province_en"]
    + EXPECTED_EN[1:]
    + ["province_translation_missing"]
]

print(clean.head())
print(ocr_header_preview)


#8) Build the review sheet and save workbook

review_cols = ["province_cn", "province_en"] + EXPECTED_EN[1:] + ["province_translation_missing"]

top_en = pd.DataFrame([{
    "province_cn": "Province (raw OCR CN)",
    "province_en": "Province (EN translation)",
    "new_capacity_total_10kw": "New capacity total",
    "new_capacity_utility_10kw": "New utility-scale",
    "new_capacity_distributed_10kw": "New distributed",
    "new_capacity_residential_10kw": "New residential",
    "cum_capacity_total_10kw": "Cumulative total",
    "cum_capacity_utility_10kw": "Cumulative utility-scale",
    "cum_capacity_distributed_10kw": "Cumulative distributed",
    "cum_capacity_residential_10kw": "Cumulative residential",
    "province_translation_missing": "Translation missing?",
}])

top_cn = pd.DataFrame([{
    "province_cn": expected_cn[0],
    "province_en": "英文省名",
    "new_capacity_total_10kw": expected_cn[1],
    "new_capacity_utility_10kw": expected_cn[2],
    "new_capacity_distributed_10kw": expected_cn[3],
    "new_capacity_residential_10kw": expected_cn[4],
    "cum_capacity_total_10kw": expected_cn[5],
    "cum_capacity_utility_10kw": expected_cn[6],
    "cum_capacity_distributed_10kw": expected_cn[7],
    "cum_capacity_residential_10kw": expected_cn[8],
    "province_translation_missing": "英文翻译缺失？",
}])

review_df = pd.concat(
    [top_en[review_cols], top_cn[review_cols], clean[review_cols]],
    ignore_index=True
)

review_xlsx = REVIEW_DIR / f"{page_slug}_review.xlsx"
clean_csv = CLEAN_CSV_DIR / f"{page_slug}_clean.csv"

with pd.ExcelWriter(review_xlsx, engine="openpyxl") as writer:
    raw.to_excel(writer, sheet_name="raw_ocr", index=False, header=False)
    ocr_header_preview.to_excel(writer, sheet_name="header_check", index=False)
    clean.to_excel(writer, sheet_name="clean", index=False)
    review_df.to_excel(writer, sheet_name="clean_review", index=False)

clean.to_csv(clean_csv, index=False, encoding="utf-8-sig")

print("saved review workbook:", review_xlsx)
print("saved clean csv:", clean_csv)


#9)  formatting for the review workbook

wb = load_workbook(review_xlsx)

ws = wb["header_check"]
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")

ws.freeze_panes = "A2"
autosize_worksheet(ws)

ws = wb["clean"]
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAD3")

ws.freeze_panes = "A2"
autosize_worksheet(ws)

ws = wb["clean_review"]
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")

for cell in ws[2]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FCE5CD")

ws.freeze_panes = "A3"
autosize_worksheet(ws)

ws = wb["raw_ocr"]
autosize_worksheet(ws, max_width=25)

ws = wb["clean"]
for col in range(3, 11):
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col).number_format = "0.0"

ws = wb["clean_review"]
for col in range(3, 11):
    for row in range(4, ws.max_row + 1):
        ws.cell(row=row, column=col).number_format = "0.0"

wb.save(review_xlsx)
print("formatted:", review_xlsx)


#10) batch loop once one page looks good

batch_items = list(enumerate(page_results))

if RERUN_ONLY_FAILED_FROM_RUN_SUMMARY:
    previous_run_summary_df = load_previous_run_summary()
    failed_urls = set(
        previous_run_summary_df.loc[
            previous_run_summary_df["status"].astype(str).str.lower().eq("failed"),
            "source_page_url",
        ].dropna()
    )
    selected_urls = failed_urls | force_rerun_urls
    batch_items = [
        (idx, page_meta)
        for idx, page_meta in enumerate(page_results)
        if page_meta["source_page_url"] in selected_urls
    ]
    run_rows = previous_run_summary_df.to_dict("records")
    print(
        "Rerun-only-failed mode:",
        f"{len(batch_items)} URLs selected from {LOGS_DIR / 'run_summary.csv'}",
    )
elif force_rerun_urls:
    batch_items = [
        (idx, page_meta)
        for idx, page_meta in enumerate(page_results)
        if page_meta["source_page_url"] in force_rerun_urls
    ]
    if (LOGS_DIR / "run_summary.csv").exists():
        run_rows = load_previous_run_summary().to_dict("records")
    else:
        run_rows = []
    print("Force-rerun mode:", f"{len(batch_items)} URLs selected")
else:
    run_rows = []

if STEP6_OCR_MODE == "openai_vision":
    n_openai_image_pages = sum(1 for _, r in batch_items if r["n_images"] > 0)
    if n_openai_image_pages > OPENAI_MAX_BATCH_IMAGE_CALLS:
        raise ValueError(
            "OpenAI batch blocked: "
            f"{n_openai_image_pages} image pages would be sent, "
            f"limit is {OPENAI_MAX_BATCH_IMAGE_CALLS}."
        )
    print(
        "OpenAI batch guard: "
        f"{n_openai_image_pages} image calls allowed "
        f"(limit {OPENAI_MAX_BATCH_IMAGE_CALLS})"
    )

for batch_pos, (PAGE_IDX, page_meta) in enumerate(batch_items, start=1):
    try:
        page_id = f"{page_meta['publish_date']}_{safe_slug(page_meta['title'])[:80]}"
        page_slug = safe_slug(page_id)
        stitched_path = STITCHED_DIR / f"{page_slug}_stitched.jpg"
        raw_table_xlsx = RAW_TABLES_DIR / f"{page_slug}_raw.xlsx"
        review_xlsx = REVIEW_DIR / f"{page_slug}_review.xlsx"
        clean_csv = CLEAN_CSV_DIR / f"{page_slug}_clean.csv"
        print(f"\n[{PAGE_IDX}] {page_meta['title']}")
        t0 = time.time()
        expected_cn = build_expected_cn_headers(page_meta["year"], page_meta["checkpoint"])
        # ============================================================
        # EXTRACT RAW TABLE
        # ============================================================
        if page_meta["n_images"] > 0:
            # image-based table -> OCR route
            raw_df = extract_raw_table_from_image(stitched_path, ocr)
            raw_df.to_excel(raw_table_xlsx, index=False, header=False)
            raw = normalize_raw_excel_frame(pd.read_excel(raw_table_xlsx, header=None))
            data_start_idx, start_col = find_total_row_and_start_col(
                raw,
                total_label=TOTAL_LABEL,
                max_search_cols=4
            )
            header_start = max(0, data_start_idx - 5)
            header_rows = raw.iloc[
                header_start:data_start_idx,
                start_col:start_col + EXPECTED_N_COLS
            ].copy()
            ocr_header_preview = pd.DataFrame({
                "col_position": range(1, EXPECTED_N_COLS + 1),
                "ocr_header_flat": [
                    flatten_header(header_rows[c].tolist()) if c in header_rows.columns else ""
                    for c in range(start_col, start_col + EXPECTED_N_COLS)
                ],
                "expected_cn": expected_cn,
                "expected_en": EXPECTED_EN,
                "friendly_en": FRIENDLY_EN,
            })
            body = raw.iloc[data_start_idx:, start_col:start_col + EXPECTED_N_COLS].copy()
            body.columns = expected_cn
        else:
            # inline HTML table -> direct text parsing route
            inline_df = extract_inline_table_from_html(page_meta["source_page_url"]).copy()
            inline_df = normalize_legacy_sparse_capacity_frame(
                inline_df,
                page_meta["year"],
            )
            if inline_df.shape[1] != EXPECTED_N_COLS:
                raise ValueError(
                    f"Inline HTML parser returned {inline_df.shape[1]} columns; expected {EXPECTED_N_COLS}"
                )
            # save a simple raw version for reference
            inline_df.to_excel(raw_table_xlsx, index=False, header=False)
            # this is what goes in the raw_ocr sheet
            raw = normalize_raw_excel_frame(inline_df.copy())
            body = raw.copy()
            body.columns = expected_cn
            # no OCR header on inline pages, so just show expected headers
            ocr_header_preview = pd.DataFrame({
                "col_position": range(1, EXPECTED_N_COLS + 1),
                "ocr_header_flat": expected_cn,
                "expected_cn": expected_cn,
                "expected_en": EXPECTED_EN,
                "friendly_en": FRIENDLY_EN,
            })
        # ============================================================
        # CLEAN
        # ============================================================
        body["省(区、市)"] = body["省(区、市)"].apply(normalize_province_name)
        body["province_en"] = body["省(区、市)"].map(PROVINCE_EN_MAP)
        for c in expected_cn[1:]:
            body[c] = body[c].apply(to_float)
        clean = body.rename(columns=dict(zip(expected_cn, EXPECTED_EN)))
        clean = repair_legacy_no_household_columns(
            clean,
            page_meta["year"],
            page_meta["checkpoint"],
        )
        clean["province_translation_missing"] = clean["province_en"].isna()
        clean = clean[
            ["province_cn", "province_en"]
            + EXPECTED_EN[1:]
            + ["province_translation_missing"]
        ]
        # ============================================================
        # BUILD REVIEW SHEET
        # ============================================================
        review_cols = ["province_cn", "province_en"] + EXPECTED_EN[1:] + ["province_translation_missing"]
        top_en = pd.DataFrame([{
            "province_cn": "Province (raw OCR CN)",
            "province_en": "Province (EN translation)",
            "new_capacity_total_10kw": "New capacity total",
            "new_capacity_utility_10kw": "New utility-scale",
            "new_capacity_distributed_10kw": "New distributed",
            "new_capacity_residential_10kw": "New residential",
            "cum_capacity_total_10kw": "Cumulative total",
            "cum_capacity_utility_10kw": "Cumulative utility-scale",
            "cum_capacity_distributed_10kw": "Cumulative distributed",
            "cum_capacity_residential_10kw": "Cumulative residential",
            "province_translation_missing": "Translation missing?",
        }])
        top_cn = pd.DataFrame([{
            "province_cn": expected_cn[0],
            "province_en": "英文省名",
            "new_capacity_total_10kw": expected_cn[1],
            "new_capacity_utility_10kw": expected_cn[2],
            "new_capacity_distributed_10kw": expected_cn[3],
            "new_capacity_residential_10kw": expected_cn[4],
            "cum_capacity_total_10kw": expected_cn[5],
            "cum_capacity_utility_10kw": expected_cn[6],
            "cum_capacity_distributed_10kw": expected_cn[7],
            "cum_capacity_residential_10kw": expected_cn[8],
            "province_translation_missing": "英文翻译缺失？",
        }])
        review_df = pd.concat(
            [top_en[review_cols], top_cn[review_cols], clean[review_cols]],
            ignore_index=True
        )
        with pd.ExcelWriter(review_xlsx, engine="openpyxl") as writer:
            raw.to_excel(writer, sheet_name="raw_ocr", index=False, header=False)
            ocr_header_preview.to_excel(writer, sheet_name="header_check", index=False)
            clean.to_excel(writer, sheet_name="clean", index=False)
            review_df.to_excel(writer, sheet_name="clean_review", index=False)
        clean.to_csv(clean_csv, index=False, encoding="utf-8-sig")
        upsert_run_row(run_rows, {
            "source_page_url": page_meta["source_page_url"],
            "title": page_meta["title"],
            "publish_date": page_meta["publish_date"],
            "year": page_meta["year"],
            "checkpoint": page_meta["checkpoint"],
            "page_type": "image_table" if page_meta["n_images"] > 0 else "inline_html_table",
            "status": "ok",
            "n_rows_clean": len(clean),
            "n_missing_translations": int(clean["province_translation_missing"].sum()),
            "elapsed_sec": round(time.time() - t0, 2),
            "notes": "",
        })
    except Exception as e:
        upsert_run_row(run_rows, {
            "source_page_url": page_meta["source_page_url"],
            "title": page_meta["title"],
            "publish_date": page_meta["publish_date"],
            "year": page_meta["year"],
            "checkpoint": page_meta["checkpoint"],
            "page_type": "image_table" if page_meta["n_images"] > 0 else "inline_html_table",
            "status": "failed",
            "n_rows_clean": pd.NA,
            "n_missing_translations": pd.NA,
            "elapsed_sec": pd.NA,
            "notes": str(e),
        })
        print("FAILED:", e)
    finally:
        run_summary_df = save_run_summary(run_rows)
        status_counts = run_summary_df["status"].value_counts().to_dict() if not run_summary_df.empty else {}
        print(
            "updated run summary:",
            LOGS_DIR / "run_summary.csv",
            f"(rerun {batch_pos}/{len(batch_items)}, summary rows {len(run_summary_df)}, {status_counts})",
        )

run_summary_df = save_run_summary(run_rows)
run_summary_df
